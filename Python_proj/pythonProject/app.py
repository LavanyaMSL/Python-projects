import openpyxl as xl
from openpyxl.chart import BarChart,Reference

def process_workbook(file_name):
    wb = xl.load_workbook(file_name)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        correct_value = cell.value * 0.9
        print(correct_value)
        correct_cell = sheet.cell(row, 4)
        correct_cell.value = correct_value

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')
    wb.save('transactions2.xlsx')

process_workbook('transactions.xlsx')
